"""
Per-protein SNP heritability for the UKB-PPP database via batched LD-score regression.

For every protein this produces two heritabilities: one on all regression variants, and one
excluding cis variants (within a window of the protein's gene, from Sun et al. 2023 ST3 hg38
gene coordinates). All proteins share one variant-index row order, so the index<->LD-score
alignment, LD scores, M and jackknife blocks are built once (PppLdscContext) and the per-
protein work is a cheap batched weighted regression (batched_ldsc_h2).

Output: a long table with two rows per protein (all_variants / cis_excluded), except proteins
absent from ST3 which get only the all_variants row.
"""

from __future__ import annotations

from pathlib import Path, PurePath

import numpy as np
import openpyxl
import polars as pl
import structlog
from attrs import frozen

from mecfs_bio.build_system.asset.base_asset import Asset
from mecfs_bio.build_system.asset.directory_asset import DirectoryAsset
from mecfs_bio.build_system.asset.file_asset import FileAsset
from mecfs_bio.build_system.meta.meta import Meta
from mecfs_bio.build_system.meta.read_spec.dataframe_read_spec import (
    DataFrameParquetFormat,
    DataFrameReadSpec,
)
from mecfs_bio.build_system.meta.read_spec.read_dataframe import scan_dataframe_asset
from mecfs_bio.build_system.meta.result_table_meta import ResultTableMeta
from mecfs_bio.build_system.rebuilder.fetch.base_fetch import Fetch
from mecfs_bio.build_system.task.base_task import GeneratingTask, Task
from mecfs_bio.build_system.task.ppp_database.build_slim_protein_parquet_task import (
    BuildSlimProteinParquetTask,
)
from mecfs_bio.build_system.task.ppp_database.protein_sample_size_task import (
    PppProteinSampleSizeTask,
)
from mecfs_bio.build_system.task.ppp_ldsc.batched_ldsc_h2 import (
    DEFAULT_N_BLOCKS,
    batched_h2,
)
from mecfs_bio.build_system.task.ppp_ldsc.ppp_ldsc_context import (
    PppLdscContext,
    build_cis_mask,
    build_ppp_ldsc_context,
    read_ld_scores,
)
from mecfs_bio.build_system.wf.base_wf import WF
from mecfs_bio.constants.gwaslab_constants import (
    GWASLAB_BETA_COL,
    GWASLAB_CHROM_COL,
    GWASLAB_POS_COL,
    GWASLAB_RSID_COL,
    GWASLAB_SE_COL,
)
from mecfs_bio.constants.ppp_index_constants import PPP_INDEX_IS_STRAND_AMBIGUOUS_COL
from mecfs_bio.constants.ppp_ldsc_constants import (
    PPP_H2_GENE_COL,
    PPP_H2_H2_COL,
    PPP_H2_H2_SE_COL,
    PPP_H2_INTERCEPT_COL,
    PPP_H2_LAMBDA_GC_COL,
    PPP_H2_MEAN_CHI2_COL,
    PPP_H2_N_BAR_COL,
    PPP_H2_N_SNPS_COL,
    PPP_H2_OID_COL,
    PPP_H2_VARIANT_SET_COL,
    PPP_N_GENE_COL,
    PPP_N_OID_COL,
    PPP_N_SAMPLE_SIZE_COL,
    PPP_N_SYNID_COL,
    PPP_VARIANT_SET_ALL,
    PPP_VARIANT_SET_CIS_EXCLUDED,
)

logger = structlog.get_logger()

_INDEX_CONTEXT_COLUMNS = [
    GWASLAB_CHROM_COL,
    GWASLAB_POS_COL,
    GWASLAB_RSID_COL,
    PPP_INDEX_IS_STRAND_AMBIGUOUS_COL,
]

# Sun et al. 2023 supplementary sheet / column labels (hg38 gene coordinates).
_ST3_SHEET = "ST3"
_ST3_OLINK_ID_COL = "Olink ID"
_ST3_GENE_CHROM_COL = "Gene CHROM"
_ST3_GENE_START_COL = "Gene start"
_ST3_GENE_END_COL = "Gene end"


@frozen
class GeneCoords:
    chrom: int
    start: int
    end: int


@frozen
class PppHeritabilityConfig:
    drop_strand_ambiguous: bool = True
    exclude_mhc: bool = True
    cis_window_bp: int = 1_000_000
    n_blocks: int = DEFAULT_N_BLOCKS
    # Proteins processed together per batched regression. Peak memory ~ n_snps * batch *
    # a few float64 arrays; 50 keeps peak comfortably within a 16 GB budget.
    batch_size: int = 50


def _parse_chrom(value: object) -> int | None:
    """Chromosome label -> int (X -> 23); None for unparseable/non-standard labels.

    NOTE:
        - There are a few "proteins" in PPP that are actually multi-protein complexes.
        - These multi-protein complexes do not correspond to a unique gene on a single chromosome.
        - In the Excel sheet, these proteins have entries in the chromosome column like chrom1;chrom2
        - We return None for these proteins.

    """
    text = str(value).strip()
    if text.isdigit():
        return int(text)
    if text.upper() == "X":
        return 23
    return None


def read_st3_gene_coords(xlsx_path: Path) -> dict[str, GeneCoords]:
    """Map Olink ID (OID) -> hg38 gene coordinates from Sun et al. 2023 supplementary ST3.

    Proteins whose chromosome label is non-standard are skipped (they then get only the
    all-variants heritability row)."""
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook[_ST3_SHEET]
    header_index: dict[str, int] | None = None
    coords: dict[str, GeneCoords] = {}
    for row in sheet.iter_rows(values_only=True):
        if header_index is None:
            if row and _ST3_OLINK_ID_COL in row:
                header_index = {
                    str(name): i for i, name in enumerate(row) if name is not None
                }
            continue
        oid = row[header_index[_ST3_OLINK_ID_COL]]
        chrom = _parse_chrom(row[header_index[_ST3_GENE_CHROM_COL]])
        start = row[header_index[_ST3_GENE_START_COL]]
        end = row[header_index[_ST3_GENE_END_COL]]
        if oid is None or chrom is None or start is None or end is None:
            continue
        coords[str(oid)] = GeneCoords(chrom=chrom, start=int(start), end=int(end))
    assert header_index is not None, f"could not find header row in sheet {_ST3_SHEET}"
    return coords


@frozen
class PppProteinHeritabilityTask(GeneratingTask):
    """Compute all-variants and cis-excluded LDSC heritability for every UKB-PPP protein.

    protein_tasks: the slim per-protein beta/se tasks (aligned to index_task).
    index_task: the shared ConstructPppVariantIndexTask (variant identity / rsID / alleles).
    ld_ref_task: an extracted reference LD-score directory (LDscore.<chr>.l2.* files).
    sample_size_task: the per-protein N table (PppProteinSampleSizeTask).
    st3_task: Sun et al. 2023 supplementary xlsx (ST3 hg38 gene coordinates).
    """

    meta: Meta
    protein_tasks: tuple[BuildSlimProteinParquetTask, ...]
    index_task: Task
    ld_ref_task: Task
    sample_size_task: PppProteinSampleSizeTask
    st3_task: Task
    config: PppHeritabilityConfig

    @property
    def deps(self) -> list[Task]:
        return [
            *self.protein_tasks,
            self.index_task,
            self.ld_ref_task,
            self.sample_size_task,
            self.st3_task,
        ]

    def execute(self, scratch_dir: Path, fetch: Fetch, wf: WF) -> Asset:
        context = self._build_context(fetch)
        sample_sizes = self._read_sample_sizes(fetch)  # synid -> (oid, gene, n)
        gene_coords = read_st3_gene_coords(self._st3_path(fetch))
        logger.info(
            "built ppp ldsc context",
            n_snps=context.n_snps,
            m=context.m,
            n_proteins=len(self.protein_tasks),
        )

        rows: list[dict] = []
        tasks = list(self.protein_tasks)
        for start in range(0, len(tasks), self.config.batch_size):
            batch = tasks[start : start + self.config.batch_size]
            rows.extend(
                self._process_batch(batch, context, sample_sizes, gene_coords, fetch)
            )

        table = pl.DataFrame(rows)
        out_path = scratch_dir / f"{self.meta.asset_id}.parquet"
        table.write_parquet(out_path)
        return FileAsset(out_path)

    def _build_context(self, fetch: Fetch) -> PppLdscContext:
        index_df = (
            scan_dataframe_asset(
                fetch(self.index_task.asset_id),
                meta=self.index_task.meta,
                parquet_backend="polars",
            )
            .to_native()
            .select(_INDEX_CONTEXT_COLUMNS)
            .collect()
        )
        ld_asset = fetch(self.ld_ref_task.asset_id)
        assert isinstance(ld_asset, DirectoryAsset)
        ld_df, m_total = read_ld_scores(ld_asset.path)
        return build_ppp_ldsc_context(
            index_df,
            ld_df,
            m_total,
            drop_strand_ambiguous=self.config.drop_strand_ambiguous,
            exclude_mhc=self.config.exclude_mhc,
        )

    def _read_sample_sizes(self, fetch: Fetch) -> dict[str, tuple[str, str, int]]:
        table = (
            scan_dataframe_asset(
                fetch(self.sample_size_task.asset_id),
                meta=self.sample_size_task.meta,
                parquet_backend="polars",
            )
            .to_native()
            .collect()
        )
        return {
            row[PPP_N_SYNID_COL]: (
                row[PPP_N_OID_COL],
                row[PPP_N_GENE_COL],
                int(row[PPP_N_SAMPLE_SIZE_COL]),
            )
            for row in table.iter_rows(named=True)
        }

    def _st3_path(self, fetch: Fetch) -> Path:
        asset = fetch(self.st3_task.asset_id)
        assert isinstance(asset, FileAsset)
        return asset.path

    def _process_batch(
        self,
        batch: list,
        context: PppLdscContext,
        sample_sizes: dict[str, tuple[str, str, int]],
        gene_coords: dict[str, GeneCoords],
        fetch: Fetch,
    ) -> list[dict]:
        oids, genes, n_vec, chi2_cols = [], [], [], []
        for protein_task in batch:
            oid, gene, n = sample_sizes[protein_task.synid]
            oids.append(oid)
            genes.append(gene)
            n_vec.append(float(n))
            chi2_cols.append(self._read_protein_chi2(protein_task, context, fetch))
        chi2 = np.column_stack(chi2_cols)
        n_arr = np.asarray(n_vec)

        cis_exclude = self._build_cis_exclude(oids, context, gene_coords)
        has_coords = [oid in gene_coords for oid in oids]

        all_res = batched_h2(
            chi2, context.ld, n_arr, context.m, n_blocks=self.config.n_blocks
        )
        cis_res = batched_h2(
            chi2,
            context.ld,
            n_arr,
            context.m,
            n_blocks=self.config.n_blocks,
            exclude=cis_exclude,
        )

        rows: list[dict] = []
        for j, oid in enumerate(oids):
            rows.append(
                self._row(oids[j], genes[j], n_arr[j], PPP_VARIANT_SET_ALL, all_res, j)
            )
            if has_coords[j]:
                rows.append(
                    self._row(
                        oids[j],
                        genes[j],
                        n_arr[j],
                        PPP_VARIANT_SET_CIS_EXCLUDED,
                        cis_res,
                        j,
                    )
                )
        return rows

    def _read_protein_chi2(
        self, protein_task: Task, context: PppLdscContext, fetch: Fetch
    ) -> np.ndarray:
        frame = (
            scan_dataframe_asset(
                fetch(protein_task.asset_id),
                meta=protein_task.meta,
                parquet_backend="polars",
            )
            .to_native()
            .select(GWASLAB_BETA_COL, GWASLAB_SE_COL)
            .collect()
        )
        beta = frame[GWASLAB_BETA_COL].to_numpy().astype(float)[context.row_pos]
        se = frame[GWASLAB_SE_COL].to_numpy().astype(float)[context.row_pos]
        with np.errstate(invalid="ignore", divide="ignore"):
            return (beta / se) ** 2

    def _build_cis_exclude(
        self,
        oids: list[str],
        context: PppLdscContext,
        gene_coords: dict[str, GeneCoords],
    ) -> np.ndarray:
        exclude = np.zeros((context.n_snps, len(oids)), dtype=bool)
        for j, oid in enumerate(oids):
            coords = gene_coords.get(oid)
            if coords is not None:
                exclude[:, j] = build_cis_mask(
                    context,
                    coords.chrom,
                    coords.start,
                    coords.end,
                    self.config.cis_window_bp,
                )
        return exclude

    @staticmethod
    def _row(oid: str, gene: str, n: float, variant_set: str, res, j: int) -> dict:
        return {
            PPP_H2_OID_COL: oid,
            PPP_H2_GENE_COL: gene,
            PPP_H2_VARIANT_SET_COL: variant_set,
            PPP_H2_H2_COL: float(res.h2[j]),
            PPP_H2_H2_SE_COL: float(res.h2_se[j]),
            PPP_H2_INTERCEPT_COL: float(res.intercept[j]),
            PPP_H2_MEAN_CHI2_COL: float(res.mean_chi2[j]),
            PPP_H2_LAMBDA_GC_COL: float(res.lambda_gc[j]),
            PPP_H2_N_SNPS_COL: int(res.n_snps[j]),
            PPP_H2_N_BAR_COL: float(n),
        }

    @classmethod
    def create(
        cls,
        asset_id: str,
        protein_tasks: tuple[BuildSlimProteinParquetTask, ...],
        index_task: Task,
        ld_ref_task: Task,
        sample_size_task: PppProteinSampleSizeTask,
        st3_task: Task,
        config: PppHeritabilityConfig = PppHeritabilityConfig(),
    ) -> PppProteinHeritabilityTask:
        meta = ResultTableMeta(
            id=asset_id,
            trait="ukbb_ppp",
            project="ppp_heritability",
            sub_dir=PurePath("analysis"),
            extension=".parquet",
            read_spec=DataFrameReadSpec(DataFrameParquetFormat()),
        )
        return cls(
            meta=meta,
            protein_tasks=protein_tasks,
            index_task=index_task,
            ld_ref_task=ld_ref_task,
            sample_size_task=sample_size_task,
            st3_task=st3_task,
            config=config,
        )
